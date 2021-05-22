# -*- coding: utf-8 -*-
"""
@Author billie
@Date 2021/4/30 11:24 上午
@Describe
    - 目标站点 https://cn163.net/archives/37436/
    - 爬取权游主要角色的信息和故事梗概
    - 存储至本地表格
@技术栈：
    - logging
    - traceback
    - requests
    - xpath
    - re
    - pyquery
    - openpyxl
    - 装饰器
    - 迭代器
"""
from functools import wraps
import traceback
import logging,os
import requests
from lxml import etree
from pyquery import PyQuery as pq
import re
import openpyxl
from openpyxl.styles import PatternFill

class Config:
    Log_path = os.path.join(os.getcwd(),'GOT.log')
    Excel_path = os.path.join(os.getcwd(),'GOT.xlsx')

logging.basicConfig(
    filename = Config.Log_path,
    level = logging.INFO,
    format = "%(asctime)s %(pathname)s[%(lineno)s]%(levelname)s:%(message)s",
    datefmt = "%Y-%m-%d %H:%M:%S %p")

def log(func):
    @wraps(func)
    def wrapper(*args,**kwargs):
        try:
            return func(*args,**kwargs)
        except Exception as e:
            logging.error(repr(e)+'\n'+traceback.format_exc())
    return wrapper

class Config:
    characters = []

class Spider:
    def __init__(self):
        self.base_url = 'https://cn163.net/archives/37436/'
        self.page = 0
        self.page_count = self.get_pageCount()

    def __str__(self):
        return "Winter is comming！"

    def __iter__(self):
        return self

    def __next__(self):
        if self.page < self.page_count:
            self.page += 1
            url = self.base_url + str(self.page)
            print(url)
            return url
    @log
    def get_pageCount(self):
        r = requests.get(self.base_url)
        eleTree = etree.HTML(r.text)
        count = eleTree.xpath('//*[@class="page-links"]/a/span/text()')[-1]
        return int(count)

    @log
    def spider(self):
        charactors = []
        en_charactors = []
        actors = []
        seasons = []
        storys = []

        r = requests.get(self.__next__())
        r.encoding = "utf-8"
        '''
        eletree = etree.HTML(r.text)
        charactors = eletree.xpath('//p/strong/span/text() | //p/span/strong/text() | //p/span/span/strong/text()')
        des = eletree.xpath('//p/span[@style = "font-size: 12pt;"]/text() | //p/span/span/text()')
        storys = eletree.xpath('//p/descendant::span[@style = "font-size: 10pt;"]/text()')
        '''
        doc = pq(r.text)
        p_s = list(doc('#post-37436 .entry-content .single-content p').items())
        p_text = [p.text() for p in p_s if not p.text()=='']
        if self.page == 1: p_text = p_text[1:]
        if self.page == 5: p_text = p_text[:-1]

        if len(p_text) > 11 or self.page == 5:
            for index, text in enumerate(p_text):
                if index % 2 == 0:
                    charactor, des = text.split('（')
                    en_charactor = re.search('(.*?)；(.*?)，(.*?)）', des).group(1)
                    actor = re.search('(.*?)；(.*?)，(.*?)）', des).group(2)
                    season = re.search('(.*?)；(.*?)，(.*?)）', des).group(3)
                    charactors.append(charactor)
                    en_charactors.append(en_charactor)
                    actors.append(actor)
                    seasons.append(season)
                elif index % 2 == 1:
                    story = text
                    storys.append(story)

        else:
            for index,text in enumerate(p_text):
                des,story = text.split('\n')

                charactor,des = des.split('（')
                en_charactor = re.search('(.*?)；(.*?)，(.*?)）',des).group(1)
                actor = re.search('(.*?)；(.*?)，(.*?)）',des).group(2)
                season = re.search('(.*?)；(.*?)，(.*?)）',des).group(3)
                charactors.append(charactor)
                en_charactors.append(en_charactor)
                actors.append(actor)
                seasons.append(season)
                storys.append(story)
        # print(len(charactors),len(en_charactors),len(actors),len(seasons),len(storys))
        for index in range(len(charactors)):
            Config.characters.append({'charactor':charactors[index],'en_charactor':en_charactors[index],'actor':actors[index],'season':seasons[index],'story':storys[index]})

    def main(self):
        while self.page < self.page_count:
            self.spider()
        logging.info(f'爬取成功，一共{len(Config.characters)}个角色')


@log
def xl_write():
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(['charactor'.upper(),'en_charactor'.upper(),'actor'.upper(),'season'.upper(),'story'.upper()])
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="ffff00")
    ws.row_dimensions[1].height = 20
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    for c in Config.characters:
        ws.append([c['charactor'],c['en_charactor'],c['actor'],c['season'],c['story']])

    wb.save(Config.Excel_path)
    logging.info('已经写入表格！')


if __name__ == '__main__':
    print(Spider())
    Spider().main()

    logging.info(Config.characters)
    for c in Config.characters: logging.info(c)
    xl_write()
